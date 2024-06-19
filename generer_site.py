

def liste_articles():
    import yaml

    with open('mkdocs.yml', 'r') as file:
        configuration = yaml.unsafe_load(file)

    liste = []
    for rubrique in configuration['nav'][2]["Leçons"]:
        for item in rubrique:
            for element in rubrique[item]:
                for k in element.keys():
                    entree = { 'titre' : k,  'fichier':element[k]}
                    liste.append(entree)
    return liste

def extraire_admonition(admonition, article):
    with open(article, 'r') as file:
        contenu = file.readlines()
    
    manuels = []
    admonition_tag = f"!!! {admonition}"

    entete_manuel = False
    for ligne in contenu:
        if admonition_tag in ligne:
            entete_manuel = True
        else:
            if entete_manuel and "](" in ligne:
                manuels.append(ligne.lstrip())
            else:
                entete_manuel = False

    return manuels

def generer_liste(admonition, entete, fichier_sortie):
    with open(fichier_sortie, 'w') as file:
        file.write(entete)

        articles = liste_articles()

        for article in articles:
            
            manuels = extraire_admonition(admonition, "wiki/" + article['fichier'])
            if manuels:
                file.write(f"## {article['titre']}")
                file.write("\n")

            for  manuel in manuels:
                file.write(manuel)

def generer_grille(fichier, nom_feuille):
    import openpyxl

    excel = openpyxl.load_workbook(fichier, data_only=True)

    feuille = excel[nom_feuille]

    padding_de_grille = '|&#8288 {: style="padding:0"}'

    grille = []

    # Écrire les entêtes du tableau de l'horaire
    entete = ""
    trait = ""
    for col in range(1, feuille.max_column + 1):
        entete += feuille.cell(row=1, column=col).value
        trait += "--"
        if col < feuille.max_column:
            entete += "|"
            trait += "|"
        else:
            entete += "\n"
            trait += "\n"

    grille.append(entete)
    grille.append(trait)

    for row in range(2, feuille.max_row + 1):
        ligne = ""
        has_colspan = False
        for col in range(1, feuille.max_column + 1):
            valeur = feuille.cell(row=row, column=col).value
            if valeur is None:
                valeur = ""
            else:
                valeur = str(valeur).replace("\n", "")
            ligne += valeur
            if col < feuille.max_column:
                if not has_colspan:
                    ligne += "|"
            else:
                if has_colspan:
                    ligne += padding_de_grille * (feuille.max_column - 1)
                ligne += "\n"
            if "colspan" in valeur:
                has_colspan = True

        

        grille.append(ligne)
    
    return grille

def generer_horaire():

    grille = generer_grille("template/horaire.xlsx", "horaire")

    with open("./wiki/horaire.md", "w") as f:
        # Écrire le titre de la page
        f.write("# Horaire du cours de piratage éthique\n")
        f.writelines(grille)


def generer_projet_integrateur():

    grille = generer_grille("template/grille_correction.xlsx", "grille")

    with open("./docs/projet_integrateur.md", "w") as f:

        with open("template/projet_integrateur.md", "r") as t:
            f.writelines(t.readlines())

        f.writelines(grille)


        
print("Générer la page de l'horaire")
generer_horaire()